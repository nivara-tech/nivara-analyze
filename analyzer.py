"""
P&L Analysis Engine for Eureka Forbes.
Reads structured P&L Excel, computes comparisons, detects outliers, generates insights.
"""
import openpyxl
import re
from dataclasses import dataclass, field
from typing import Optional


FY_MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
QUARTERS = {
    "Q1": ["Apr", "May", "Jun"],
    "Q2": ["Jul", "Aug", "Sep"],
    "Q3": ["Oct", "Nov", "Dec"],
    "Q4": ["Jan", "Feb", "Mar"],
}

# Threshold for flagging outliers (absolute % deviation)
OUTLIER_THRESHOLD_PCT = 10  # flag if >10% deviation


@dataclass
class MonthData:
    label: str
    unit: str
    actual: Optional[float] = None
    aop: Optional[float] = None
    prev_month_actual: Optional[float] = None
    ly_same_month_actual: Optional[float] = None


@dataclass
class Outlier:
    line_item: str
    comparison_type: str  # "vs_AOP", "MoM", "YoY", "QoQ"
    current_value: float
    reference_value: float
    deviation_pct: float
    direction: str  # "above" or "below"
    severity: str  # "high", "medium"

    @property
    def description(self):
        ref_label = {
            "vs_AOP": "AOP/Budget",
            "MoM": "Previous Month",
            "YoY": "Last Year Same Month",
            "QoQ": "Last Quarter Avg",
        }[self.comparison_type]
        dir_word = "above" if self.direction == "above" else "below"
        return (
            f"{self.line_item}: {abs(self.deviation_pct):.1f}% {dir_word} {ref_label} "
            f"(Actual: {self.current_value:.1f} vs {ref_label}: {self.reference_value:.1f})"
        )


@dataclass
class AnalysisResult:
    company: str
    review_month: str
    review_fy: str
    review_fy_int: int

    # Current month data
    current_month: dict = field(default_factory=dict)  # label -> value
    aop_month: dict = field(default_factory=dict)
    prev_month: dict = field(default_factory=dict)
    ly_same_month: dict = field(default_factory=dict)

    # Quarter data
    current_quarter: dict = field(default_factory=dict)
    aop_quarter: dict = field(default_factory=dict)
    ly_same_quarter: dict = field(default_factory=dict)
    prev_quarter: dict = field(default_factory=dict)

    # Full year (YTD and full year for March)
    ytd_actual: dict = field(default_factory=dict)
    ytd_aop: dict = field(default_factory=dict)
    ly_full_year: dict = field(default_factory=dict)
    fy_aop: dict = field(default_factory=dict)

    # Outliers
    outliers: list = field(default_factory=list)

    # Line items and units
    line_items: list = field(default_factory=list)
    units: dict = field(default_factory=dict)

    # Key highlights (auto-generated)
    highlights: list = field(default_factory=list)

    # Budget achievement
    budget_achievement: dict = field(default_factory=dict)

    # Monthly trend (all months current FY)
    monthly_trend: dict = field(default_factory=dict)  # label -> [values]


def _parse_col_header(header):
    """Parse column header like 'Mar FY26 (A)' -> (month, fy, type)"""
    match = re.match(r'(\w+)\s+FY(\d+)\s+\((\w+)\)', str(header))
    if match:
        return match.group(1), int(match.group(2)), match.group(3)
    return None, None, None


def _get_quarter(month_name):
    for q, months in QUARTERS.items():
        if month_name in months:
            return q
    return None


def _prev_quarter(quarter):
    qs = ["Q1", "Q2", "Q3", "Q4"]
    idx = qs.index(quarter)
    return qs[idx - 1] if idx > 0 else "Q4"


def _safe_pct(a, b):
    if a is None or b is None:
        return 0
    if b != 0:
        return ((a - b) / abs(b)) * 100
    return 0


def _aggregate_months(data_by_col, months, fy, col_type, line_items):
    """Sum monthly values for given months."""
    result = {}
    for label in line_items:
        total = 0
        count = 0
        for m in months:
            key = f"{m}_FY{fy}_{col_type}"
            if key in data_by_col and label in data_by_col[key]:
                val = data_by_col[key][label]
                if val is not None and not isinstance(val, str):
                    total += val
                    count += 1
        result[label] = total if count > 0 else None
    return result


def analyze_pnl(excel_path, review_month=None, review_fy=None):
    """
    Main analysis function. Reads Excel and produces full analysis.

    Args:
        excel_path: Path to the P&L Excel file
        review_month: Month to review (e.g., "Mar"). Auto-detected from Metadata if None.
        review_fy: FY year (e.g., 2026). Auto-detected from Metadata if None.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Try to read metadata
    if "Metadata" in wb.sheetnames:
        meta = wb["Metadata"]
        if not review_month:
            review_month = meta["B2"].value
        if not review_fy:
            fy_str = str(meta["B3"].value)
            review_fy = 2000 + int(fy_str.replace("FY", ""))

    ws = wb["P&L Data"]

    # Parse headers
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=1, column=col).value)

    # Parse column mapping
    col_map = {}  # col_index -> (month, fy, type)
    for ci, h in enumerate(headers):
        if h and ci >= 2:
            m, fy, typ = _parse_col_header(h)
            if m:
                key = f"{m}_FY{fy}_{typ}"
                col_map[ci] = key

    # Read all data
    line_items = []
    units = {}
    data_by_col = {}  # key -> {label: value}

    for key in col_map.values():
        data_by_col[key] = {}

    for row in range(2, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        unit = ws.cell(row=row, column=2).value
        if not label:
            continue
        line_items.append(label)
        units[label] = unit or ""

        for ci, key in col_map.items():
            val = ws.cell(row=row, column=ci + 1).value
            if val is not None and not isinstance(val, str):
                data_by_col[key][label] = val

    # Determine review parameters
    fy_short = review_fy % 100
    prev_fy_short = fy_short - 1
    month_idx = FY_MONTHS.index(review_month)
    prev_month = FY_MONTHS[month_idx - 1] if month_idx > 0 else "Mar"
    prev_month_fy = fy_short if month_idx > 0 else prev_fy_short

    current_quarter = _get_quarter(review_month)
    current_q_months = QUARTERS[current_quarter]
    # Only include months up to and including review month within the quarter
    q_month_idx_in_q = current_q_months.index(review_month)
    active_q_months = current_q_months[:q_month_idx_in_q + 1]

    prev_q = _prev_quarter(current_quarter)
    prev_q_fy = fy_short if current_quarter != "Q1" else prev_fy_short

    # Current month data
    cm_key = f"{review_month}_FY{fy_short}_A"
    aop_key = f"{review_month}_FY{fy_short}_AOP"
    pm_key = f"{prev_month}_FY{prev_month_fy}_A"
    ly_key = f"{review_month}_FY{prev_fy_short}_A"

    current_month = data_by_col.get(cm_key, {})
    aop_month = data_by_col.get(aop_key, {})
    prev_month_data = data_by_col.get(pm_key, {})
    ly_same_month = data_by_col.get(ly_key, {})

    # Quarter aggregation
    current_quarter_data = _aggregate_months(data_by_col, active_q_months, fy_short, "A", line_items)
    aop_quarter_data = _aggregate_months(data_by_col, active_q_months, fy_short, "AOP", line_items)
    ly_quarter_data = _aggregate_months(data_by_col, active_q_months, prev_fy_short, "A", line_items)

    # Previous quarter (full quarter)
    prev_quarter_data = _aggregate_months(data_by_col, QUARTERS[prev_q], prev_q_fy, "A", line_items)

    # YTD (Apr to review month)
    ytd_months = FY_MONTHS[:month_idx + 1]
    ytd_actual = _aggregate_months(data_by_col, ytd_months, fy_short, "A", line_items)
    ytd_aop = _aggregate_months(data_by_col, ytd_months, fy_short, "AOP", line_items)

    # Last year full year
    ly_full_year = _aggregate_months(data_by_col, FY_MONTHS, prev_fy_short, "A", line_items)

    # Full year AOP
    fy_aop = _aggregate_months(data_by_col, FY_MONTHS, fy_short, "AOP", line_items)

    # Monthly trend for current FY
    monthly_trend = {}
    for label in line_items:
        monthly_trend[label] = []
        for m in FY_MONTHS[:month_idx + 1]:
            k = f"{m}_FY{fy_short}_A"
            val = data_by_col.get(k, {}).get(label)
            monthly_trend[label].append(val if val is not None else 0)

    # Detect outliers
    outliers = []
    key_lines = [li for li in line_items if units.get(li) == "Rs Cr"]

    for label in key_lines:
        cv = current_month.get(label)
        if cv is None or cv == 0:
            continue

        # vs AOP
        av = aop_month.get(label)
        if av and av != 0:
            dev = _safe_pct(cv, av)
            if abs(dev) > OUTLIER_THRESHOLD_PCT:
                outliers.append(Outlier(
                    line_item=label, comparison_type="vs_AOP",
                    current_value=cv, reference_value=av,
                    deviation_pct=dev,
                    direction="above" if dev > 0 else "below",
                    severity="high" if abs(dev) > 20 else "medium"
                ))

        # MoM
        pv = prev_month_data.get(label)
        if pv and pv != 0:
            dev = _safe_pct(cv, pv)
            if abs(dev) > OUTLIER_THRESHOLD_PCT:
                outliers.append(Outlier(
                    line_item=label, comparison_type="MoM",
                    current_value=cv, reference_value=pv,
                    deviation_pct=dev,
                    direction="above" if dev > 0 else "below",
                    severity="high" if abs(dev) > 20 else "medium"
                ))

        # YoY
        lv = ly_same_month.get(label)
        if lv and lv != 0:
            dev = _safe_pct(cv, lv)
            if abs(dev) > OUTLIER_THRESHOLD_PCT:
                outliers.append(Outlier(
                    line_item=label, comparison_type="YoY",
                    current_value=cv, reference_value=lv,
                    deviation_pct=dev,
                    direction="above" if dev > 0 else "below",
                    severity="high" if abs(dev) > 20 else "medium"
                ))

    # QoQ outliers (current quarter vs previous quarter average per month)
    for label in key_lines:
        cq = current_quarter_data.get(label)
        pq = prev_quarter_data.get(label)
        if cq and pq and pq != 0:
            # Normalize by number of months
            cq_avg = cq / len(active_q_months) if active_q_months else 0
            pq_avg = pq / len(QUARTERS[prev_q])
            if pq_avg != 0:
                dev = _safe_pct(cq_avg, pq_avg)
                if abs(dev) > OUTLIER_THRESHOLD_PCT:
                    outliers.append(Outlier(
                        line_item=label, comparison_type="QoQ",
                        current_value=cq_avg, reference_value=pq_avg,
                        deviation_pct=dev,
                        direction="above" if dev > 0 else "below",
                        severity="high" if abs(dev) > 20 else "medium"
                    ))

    # Sort outliers by severity then absolute deviation
    outliers.sort(key=lambda o: (-1 if o.severity == "high" else 0, -abs(o.deviation_pct)))

    # Generate key highlights
    highlights = []

    # Budget achievement for key metrics
    budget_achievement = {}
    key_metrics = ["Total Net Sales", "Gross Margin", "EBITDA (post allocation)",
                   "Profit Before Tax", "Profit After Tax", "Total COGS",
                   "Total Direct Costs", "Freight", "Total Employee Related Costs",
                   "Total Adv & Sales Promo"]

    for label in key_metrics:
        actual = current_month.get(label, 0)
        budget = aop_month.get(label, 0)
        if budget and budget != 0:
            ach_pct = (actual / budget) * 100
            budget_achievement[label] = {
                "actual": actual,
                "budget": budget,
                "achievement_pct": ach_pct,
                "variance": actual - budget,
                "variance_pct": _safe_pct(actual, budget),
            }

    # Auto highlights
    # Revenue
    ns_ach = budget_achievement.get("Total Net Sales", {})
    if ns_ach:
        highlights.append(f"Revenue AOP Achievement: {ns_ach['achievement_pct']:.0f}% (Rs {ns_ach['actual']:.1f} Cr vs AOP Rs {ns_ach['budget']:.1f} Cr)")

    # EBITDA
    eb_ach = budget_achievement.get("EBITDA (post allocation)", {})
    if eb_ach:
        highlights.append(f"EBITDA AOP Achievement: {eb_ach['achievement_pct']:.0f}% (Rs {eb_ach['actual']:.1f} Cr vs AOP Rs {eb_ach['budget']:.1f} Cr)")

    # Freight outlier
    freight_outliers = [o for o in outliers if "Freight" in o.line_item]
    for fo in freight_outliers:
        highlights.append(f"Freight cost {fo.description}")

    # Top 3 cost outliers (costs going up = bad)
    cost_outliers = [o for o in outliers if o.direction == "above" and
                     any(kw in o.line_item for kw in ["Cost", "Expense", "Freight", "Commission", "Charges"])]
    for co in cost_outliers[:3]:
        if co.line_item not in [h.split(":")[0] for h in highlights]:
            highlights.append(co.description)

    # YoY growth
    ns_yoy_actual = current_month.get("Total Net Sales", 0)
    ns_yoy_ly = ly_same_month.get("Total Net Sales", 0)
    if ns_yoy_ly and ns_yoy_ly != 0:
        yoy_g = _safe_pct(ns_yoy_actual, ns_yoy_ly)
        highlights.append(f"YoY Revenue Growth: {yoy_g:.1f}% (Rs {ns_yoy_actual:.1f} Cr vs LY Rs {ns_yoy_ly:.1f} Cr)")

    # Full year view for March
    is_full_year = review_month == "Mar"

    result = AnalysisResult(
        company="Eureka Forbes Limited",
        review_month=review_month,
        review_fy=f"FY{fy_short}",
        review_fy_int=review_fy,
        current_month=current_month,
        aop_month=aop_month,
        prev_month=prev_month_data,
        ly_same_month=ly_same_month,
        current_quarter=current_quarter_data,
        aop_quarter=aop_quarter_data,
        ly_same_quarter=ly_quarter_data,
        prev_quarter=prev_quarter_data,
        ytd_actual=ytd_actual,
        ytd_aop=ytd_aop,
        ly_full_year=ly_full_year,
        fy_aop=fy_aop,
        outliers=outliers,
        line_items=line_items,
        units=units,
        highlights=highlights,
        budget_achievement=budget_achievement,
        monthly_trend=monthly_trend,
    )

    return result
