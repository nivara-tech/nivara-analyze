"""
Generate realistic sample P&L data for Eureka Forbes in the same structure as the real file.
Creates an Excel file with monthly actuals, AOP (budget), and historical data.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
import datetime

# P&L line items matching the real Eureka Forbes structure
PNL_LINES = [
    # (row_label, unit, category, base_monthly_value_cr)
    # Revenue section - Product breakup
    ("Gross Sales - EWP", "Rs Cr", "revenue_product", 165),
    ("Gross Sales - VC", "Rs Cr", "revenue_product", 55),
    ("Gross Sales - NEWP", "Rs Cr", "revenue_product", 12),
    ("Gross Sales - Air Purifier", "Rs Cr", "revenue_product", 18),
    ("Gross Sales - Softeners", "Rs Cr", "revenue_product", 15),
    ("Gross Sales - Others", "Rs Cr", "revenue_product", 15),
    ("Total Product Gross Sales", "Rs Cr", "total_product_gs", 280),

    # Net Sales
    ("Net Sales - Products", "Rs Cr", "revenue", 280),
    ("Net Sales - AMC", "Rs Cr", "revenue", 120),
    ("Net Sales - Product Rental", "Rs Cr", "revenue", 15),
    ("Net Sales - Sale of Spares", "Rs Cr", "revenue", 25),
    ("Net Sales - Service Charge", "Rs Cr", "revenue", 18),
    ("Total Net Sales", "Rs Cr", "total_revenue", 458),

    # COGS
    ("Product COGS", "Rs Cr", "cogs", 112),
    ("Spare & Consumable COGS", "Rs Cr", "cogs", 18),
    ("Provision for COGS", "Rs Cr", "cogs", 3),
    ("Total COGS", "Rs Cr", "total_cogs", 133),

    # Gross Margin
    ("Gross Margin", "Rs Cr", "gross_margin", 325),
    ("Gross Margin %", "%", "margin_pct", 70.9),

    # Direct Costs
    ("Employee Cost (excl Incentives)", "Rs Cr", "direct_cost", 52),
    ("Incentives", "Rs Cr", "direct_cost", 18),
    ("Employee Cost - Total", "Rs Cr", "direct_cost", 70),
    ("SC Cost", "Rs Cr", "direct_cost", 12),
    ("Staff Welfare", "Rs Cr", "direct_cost", 3),
    ("Vehicle", "Rs Cr", "direct_cost", 8),
    ("Total Employee Related Costs", "Rs Cr", "total_emp", 93),

    # Advertisement & Sales Promotion
    ("ATL (Digital, Non-digital, Prod Comm)", "Rs Cr", "adv_cost", 15),
    ("Digital Performance MKT & Others", "Rs Cr", "adv_cost", 8),
    ("Total Advertisement Costs", "Rs Cr", "adv_cost", 23),
    ("Total Sales Promotion", "Rs Cr", "adv_cost", 12),
    ("Total Adv & Sales Promo", "Rs Cr", "total_adv", 35),

    # Service Charges
    ("Call Related Charges", "Rs Cr", "svc_cost", 4),
    ("ASC Spare Claim", "Rs Cr", "svc_cost", 6),
    ("AMC Commission", "Rs Cr", "svc_cost", 8),
    ("Product Sale Commission", "Rs Cr", "svc_cost", 5),
    ("BP & ST Incentive", "Rs Cr", "svc_cost", 3),
    ("Total Service Charges", "Rs Cr", "total_svc", 26),

    # Other Direct Costs
    ("Freight", "Rs Cr", "other_direct", 14),
    ("Rates and Taxes", "Rs Cr", "other_direct", 2),
    ("Logistics Expenses", "Rs Cr", "other_direct", 5),
    ("Conference", "Rs Cr", "other_direct", 1.5),
    ("Bad Debts", "Rs Cr", "other_direct", 1),
    ("Total Other Direct Costs", "Rs Cr", "total_other_direct", 23.5),

    ("Total Direct Costs", "Rs Cr", "total_direct", 177.5),
    ("Contribution Profit (GM-DC)", "Rs Cr", "contribution", 147.5),
    ("Contribution Profit %", "%", "margin_pct", 32.2),

    # Indirect Costs
    ("Information Technology", "Rs Cr", "indirect", 6),
    ("Legal and Professional", "Rs Cr", "indirect", 3),
    ("Training Cost", "Rs Cr", "indirect", 1),
    ("Wages to Contractual", "Rs Cr", "indirect", 4),
    ("R&D", "Rs Cr", "indirect", 5),
    ("Bank Charges", "Rs Cr", "indirect", 1.5),
    ("Sundry Expenses", "Rs Cr", "indirect", 2),
    ("Travel", "Rs Cr", "indirect", 4),
    ("Communication", "Rs Cr", "indirect", 1.5),
    ("Total Indirect Costs (A)", "Rs Cr", "total_indirect_a", 28),

    ("Rent", "Rs Cr", "indirect_b", 8),
    ("Repairs", "Rs Cr", "indirect_b", 2),
    ("Insurance", "Rs Cr", "indirect_b", 1),
    ("Electricity", "Rs Cr", "indirect_b", 1.5),
    ("SP Brand Fees", "Rs Cr", "indirect_b", 3),
    ("Total Indirect Costs (B)", "Rs Cr", "total_indirect_b", 15.5),

    ("Total Indirect Costs (A+B)", "Rs Cr", "total_indirect", 43.5),

    # Operating metrics
    ("Other Operating Income", "Rs Cr", "other_income", 5),
    ("EBITDA Operating", "Rs Cr", "ebitda_op", 109),
    ("EBITDA Operating %", "%", "margin_pct", 23.8),

    ("Common/Corp OH", "Rs Cr", "overhead", 12),
    ("Mfg Overheads", "Rs Cr", "overhead", 8),
    ("Common + Mfg Overheads", "Rs Cr", "total_overhead", 20),

    ("EBITDA (post allocation)", "Rs Cr", "ebitda", 89),
    ("EBITDA %", "%", "margin_pct", 19.4),

    ("Depreciation", "Rs Cr", "dep", 12),
    ("EBIT", "Rs Cr", "ebit", 77),
    ("EBIT %", "%", "margin_pct", 16.8),

    ("Other Non-Operating Income", "Rs Cr", "non_op", 2),
    ("Interest Cost", "Rs Cr", "interest", 4),
    ("Profit Before Tax", "Rs Cr", "pbt", 75),
    ("PBT %", "%", "margin_pct", 16.4),

    ("Tax", "Rs Cr", "tax", 19),
    ("Profit After Tax", "Rs Cr", "pat", 56),
    ("PAT %", "%", "margin_pct", 12.2),
]

# Months in Indian FY (Apr-Mar)
FY_MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# Seasonality factors (water purifier business is seasonal)
SEASONALITY = {
    "Apr": 1.15, "May": 1.20, "Jun": 1.10, "Jul": 0.95,
    "Aug": 0.90, "Sep": 0.85, "Oct": 0.95, "Nov": 1.00,
    "Dec": 0.92, "Jan": 0.88, "Feb": 0.95, "Mar": 1.15
}


def _add_noise(val, pct=0.05):
    return val * (1 + random.uniform(-pct, pct))


def _compute_totals(row_data):
    """Recompute total/derived rows from component rows."""
    # Map label -> value for this column
    by_label = {label: val for label, val in row_data.items()}

    # Total Product Gross Sales
    by_label["Total Product Gross Sales"] = sum(by_label.get(k, 0) for k in [
        "Gross Sales - EWP", "Gross Sales - VC", "Gross Sales - NEWP",
        "Gross Sales - Air Purifier", "Gross Sales - Softeners", "Gross Sales - Others"
    ])

    # Total Net Sales
    by_label["Total Net Sales"] = sum(by_label.get(k, 0) for k in [
        "Net Sales - Products", "Net Sales - AMC", "Net Sales - Product Rental",
        "Net Sales - Sale of Spares", "Net Sales - Service Charge"
    ])

    # Total COGS
    by_label["Total COGS"] = sum(by_label.get(k, 0) for k in [
        "Product COGS", "Spare & Consumable COGS", "Provision for COGS"
    ])

    ns = by_label["Total Net Sales"]

    # Gross Margin
    by_label["Gross Margin"] = ns - by_label["Total COGS"]
    by_label["Gross Margin %"] = (by_label["Gross Margin"] / ns * 100) if ns else 0

    # Employee totals
    by_label["Employee Cost - Total"] = by_label.get("Employee Cost (excl Incentives)", 0) + by_label.get("Incentives", 0)
    by_label["Total Employee Related Costs"] = (
        by_label["Employee Cost - Total"] + by_label.get("SC Cost", 0) +
        by_label.get("Staff Welfare", 0) + by_label.get("Vehicle", 0)
    )

    # Ad totals
    by_label["Total Advertisement Costs"] = by_label.get("ATL (Digital, Non-digital, Prod Comm)", 0) + by_label.get("Digital Performance MKT & Others", 0)
    by_label["Total Adv & Sales Promo"] = by_label["Total Advertisement Costs"] + by_label.get("Total Sales Promotion", 0)

    # Service charges total
    by_label["Total Service Charges"] = sum(by_label.get(k, 0) for k in [
        "Call Related Charges", "ASC Spare Claim", "AMC Commission",
        "Product Sale Commission", "BP & ST Incentive"
    ])

    # Other direct
    by_label["Total Other Direct Costs"] = sum(by_label.get(k, 0) for k in [
        "Freight", "Rates and Taxes", "Logistics Expenses", "Conference", "Bad Debts"
    ])

    # Total Direct
    by_label["Total Direct Costs"] = (
        by_label["Total Employee Related Costs"] + by_label["Total Adv & Sales Promo"] +
        by_label["Total Service Charges"] + by_label["Total Other Direct Costs"]
    )

    # Contribution
    by_label["Contribution Profit (GM-DC)"] = by_label["Gross Margin"] - by_label["Total Direct Costs"]
    by_label["Contribution Profit %"] = (by_label["Contribution Profit (GM-DC)"] / ns * 100) if ns else 0

    # Indirect A
    by_label["Total Indirect Costs (A)"] = sum(by_label.get(k, 0) for k in [
        "Information Technology", "Legal and Professional", "Training Cost",
        "Wages to Contractual", "R&D", "Bank Charges", "Sundry Expenses",
        "Travel", "Communication"
    ])

    # Indirect B
    by_label["Total Indirect Costs (B)"] = sum(by_label.get(k, 0) for k in [
        "Rent", "Repairs", "Insurance", "Electricity", "SP Brand Fees"
    ])

    by_label["Total Indirect Costs (A+B)"] = by_label["Total Indirect Costs (A)"] + by_label["Total Indirect Costs (B)"]

    # EBITDA
    by_label["EBITDA Operating"] = by_label["Contribution Profit (GM-DC)"] - by_label["Total Indirect Costs (A+B)"] + by_label.get("Other Operating Income", 0)
    by_label["EBITDA Operating %"] = (by_label["EBITDA Operating"] / ns * 100) if ns else 0

    by_label["Common + Mfg Overheads"] = by_label.get("Common/Corp OH", 0) + by_label.get("Mfg Overheads", 0)

    by_label["EBITDA (post allocation)"] = by_label["EBITDA Operating"] - by_label["Common + Mfg Overheads"]
    by_label["EBITDA %"] = (by_label["EBITDA (post allocation)"] / ns * 100) if ns else 0

    by_label["EBIT"] = by_label["EBITDA (post allocation)"] - by_label.get("Depreciation", 0)
    by_label["EBIT %"] = (by_label["EBIT"] / ns * 100) if ns else 0

    by_label["Profit Before Tax"] = by_label["EBIT"] + by_label.get("Other Non-Operating Income", 0) - by_label.get("Interest Cost", 0)
    by_label["PBT %"] = (by_label["Profit Before Tax"] / ns * 100) if ns else 0

    by_label["Tax"] = by_label["Profit Before Tax"] * 0.252  # ~25.2% effective tax
    by_label["Profit After Tax"] = by_label["Profit Before Tax"] - by_label["Tax"]
    by_label["PAT %"] = (by_label["Profit After Tax"] / ns * 100) if ns else 0

    return by_label


def generate_monthly_data(base_lines, fy_year, month_name, yoy_growth=0.08, is_aop=False):
    """Generate one month of P&L data."""
    season = SEASONALITY[month_name]
    data = {}

    for label, unit, category, base_val in base_lines:
        if category in ("margin_pct",):
            continue  # computed

        if category in ("total_revenue", "total_cogs", "total_emp", "total_adv",
                        "total_svc", "total_other_direct", "total_direct",
                        "contribution", "total_indirect_a", "total_indirect_b",
                        "total_indirect", "ebitda_op", "total_overhead", "ebitda",
                        "ebit", "pbt", "pat"):
            continue  # computed from components

        # Apply growth year-over-year
        growth_factor = 1 + yoy_growth * (fy_year - 2024)

        if is_aop:
            val = base_val * growth_factor * season * random.uniform(0.97, 1.03)
        else:
            # Actuals have more variance - sometimes beat, sometimes miss budget
            val = base_val * growth_factor * season * random.uniform(0.88, 1.12)

        data[label] = round(val, 2)

    return _compute_totals(data)


def generate_sample_excel(output_path, review_month="Mar", review_fy=2026):
    """
    Generate a sample P&L Excel file structured for analysis.

    Columns: Line Item | Unit | FY25 Months (Actuals) | FY26 Months (Actuals up to review month, AOP after) | AOP FY26
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Data"

    random.seed(42)  # reproducible

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    pct_font = Font(italic=True, size=9, color="666666")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Determine which months have actuals vs forecast
    month_idx = FY_MONTHS.index(review_month)

    # Build column headers
    headers = ["Line Item", "Unit"]

    # Previous year (FY25) - all 12 months actuals
    prev_fy = review_fy - 1
    for m in FY_MONTHS:
        headers.append(f"{m} FY{prev_fy % 100} (A)")

    # Current year (FY26) - actuals up to review month, then blank
    for i, m in enumerate(FY_MONTHS):
        if i <= month_idx:
            headers.append(f"{m} FY{review_fy % 100} (A)")
        else:
            headers.append(f"{m} FY{review_fy % 100} (F)")

    # AOP for current year (full year budget)
    for m in FY_MONTHS:
        headers.append(f"{m} FY{review_fy % 100} (AOP)")

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Generate data for each line item
    all_data = {}

    # Previous year actuals
    for mi, m in enumerate(FY_MONTHS):
        key = f"{m}_FY{prev_fy % 100}_A"
        all_data[key] = generate_monthly_data(PNL_LINES, prev_fy, m, yoy_growth=0.08, is_aop=False)

    # Current year actuals (up to review month)
    for i, m in enumerate(FY_MONTHS):
        if i <= month_idx:
            key = f"{m}_FY{review_fy % 100}_A"
            all_data[key] = generate_monthly_data(PNL_LINES, review_fy, m, yoy_growth=0.10, is_aop=False)
        else:
            # Forecast for remaining months (close to AOP)
            key = f"{m}_FY{review_fy % 100}_F"
            all_data[key] = generate_monthly_data(PNL_LINES, review_fy, m, yoy_growth=0.10, is_aop=True)

    # AOP (budget) for current year
    for m in FY_MONTHS:
        key = f"{m}_FY{review_fy % 100}_AOP"
        all_data[key] = generate_monthly_data(PNL_LINES, review_fy, m, yoy_growth=0.10, is_aop=True)

    # Build column keys in order
    col_keys = []
    for m in FY_MONTHS:
        col_keys.append(f"{m}_FY{prev_fy % 100}_A")
    for i, m in enumerate(FY_MONTHS):
        if i <= month_idx:
            col_keys.append(f"{m}_FY{review_fy % 100}_A")
        else:
            col_keys.append(f"{m}_FY{review_fy % 100}_F")
    for m in FY_MONTHS:
        col_keys.append(f"{m}_FY{review_fy % 100}_AOP")

    # Write line items
    row_num = 2
    for label, unit, category, base_val in PNL_LINES:
        ws.cell(row=row_num, column=1, value=label).border = thin_border
        ws.cell(row=row_num, column=2, value=unit).border = thin_border

        is_total = category.startswith("total_") or category in ("gross_margin", "contribution", "ebitda_op", "ebitda", "ebit", "pbt", "pat")
        is_pct = category == "margin_pct"

        for ci, ck in enumerate(col_keys):
            val = all_data[ck].get(label, 0)
            cell = ws.cell(row=row_num, column=3 + ci, value=round(val, 1))
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')

            if is_pct:
                cell.number_format = '0.0"%"'
                cell.font = pct_font
            else:
                cell.number_format = '#,##0.0'

            if is_total:
                cell.font = total_font
                cell.fill = total_fill

        row_num += 1

    # Auto-width columns
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 8
    for col in range(3, 3 + len(col_keys)):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Freeze panes
    ws.freeze_panes = "C2"

    # Add a metadata sheet
    ws2 = wb.create_sheet("Metadata")
    ws2["A1"] = "Company"
    ws2["B1"] = "Eureka Forbes Limited"
    ws2["A2"] = "Review Month"
    ws2["B2"] = review_month
    ws2["A3"] = "Review FY"
    ws2["B3"] = f"FY{review_fy % 100}"
    ws2["A4"] = "Generated"
    ws2["B4"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A5"] = "Currency"
    ws2["B5"] = "Rs Crores"

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = generate_sample_excel("sample_pnl_eureka_forbes.xlsx", review_month="Mar", review_fy=2026)
    print(f"Sample P&L generated: {path}")
